import openpyxl as xl
import pdfplumber
import os

#buscar arquivos pdf

for arquivo in os.listdir('files'):

    if arquivo.lower().endswith('.pdf'):
        try:
            # abrindo o arquivo Excel
            excel = xl.load_workbook("Base de Dados Inspeções.xlsx")
            aba = excel.active
        
            row_inicio = len(aba["A"])+1
            # ler o arquivo pdf e extrair os dados
            pdf = pdfplumber.open(f'files\\{arquivo}')
            pagina = pdf.pages[0]

            dados = pagina.extract_table()

            for indice, dado in enumerate(dados[1:], start=row_inicio):
                
                if dado[0] == "":
                    pass
               
                else:
                    aba.cell(row=indice, column=1).value = dado[0]
                    aba.cell(row=indice, column=2).value = dado[1]
                    aba.cell(row=indice, column=3).value = dado[2]
                    aba.cell(row=indice, column=4).value = dado[3]
                    aba.cell(row=indice, column=5).value = dado[4]

                    

            pdf.close()
            excel.save("Base de Dados Inspeções.xlsx")
            excel.close()

        except Exception as e:
            with open('log_erros.txt', 'a') as log:
                log.write(f'Houve um erro ao extrair informações do arquivo {arquivo} \n')
                log.write(f'Erro: {e}')

    else:
        with open('log_erros.txt', 'a') as log:
            log.write(f'O arquivo {arquivo} não é um PDF válido! \n')